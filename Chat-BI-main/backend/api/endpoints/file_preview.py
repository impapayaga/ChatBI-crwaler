"""
文件预览API端点
提供MinIO中原文件的内容预览功能
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import Optional
import logging
import io
import pandas as pd
import json

from models.sys_dataset import SysDataset
from core.minio_client import minio_client
from api.dependencies.dependencies import get_async_session

router = APIRouter()
logger = logging.getLogger(__name__)


@router.get("/datasets/{dataset_id}/original-file-preview")
async def get_original_file_preview(
    dataset_id: str,
    lines: int = 100,
    session: AsyncSession = Depends(get_async_session)
):
    """
    获取数据集原文件的预览内容
    
    Args:
        dataset_id: 数据集ID
        lines: 预览行数，默认100行
        session: 数据库会话
        
    Returns:
        原文件的预览内容
    """
    try:
        # 查询数据集信息
        result = await session.execute(
            select(SysDataset).where(SysDataset.id == dataset_id)
        )
        dataset = result.scalar_one_or_none()
        
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
            
        if not dataset.original_file_path:
            raise HTTPException(status_code=404, detail="原文件路径不存在")
            
        # 从MinIO路径中提取对象名称
        # 例如: chatbi-datasets/uploads/xxx_file.csv -> uploads/xxx_file.csv
        object_name = dataset.original_file_path
        if '/' in object_name and object_name.count('/') > 1:
            # 如果包含bucket名称，去掉bucket部分
            parts = object_name.split('/', 1)
            if len(parts) > 1:
                object_name = parts[1]
        
        logger.info(f"正在预览文件: {object_name}")
        
        # 从MinIO下载文件
        file_data = minio_client.download_file(object_name)
        
        # 根据文件扩展名解析内容
        file_name = dataset.name.lower()
        
        if file_name.endswith('.csv'):
            # CSV文件预览
            try:
                # 尝试不同的编码
                encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
                df = None
                used_encoding = None
                
                for encoding in encodings:
                    try:
                        df = pd.read_csv(
                            io.BytesIO(file_data),
                            encoding=encoding,
                            nrows=lines
                        )
                        used_encoding = encoding
                        break
                    except UnicodeDecodeError:
                        continue
                
                if df is None:
                    raise HTTPException(status_code=500, detail="无法解析CSV文件编码")
                
                # 转换为JSON格式返回
                preview_data = df.to_dict('records')
                
                return {
                    "success": True,
                    "file_type": "csv",
                    "encoding": used_encoding,
                    "total_rows": len(preview_data),
                    "columns": list(df.columns),
                    "data": preview_data,
                    "message": f"显示前 {len(preview_data)} 行数据"
                }
                
            except Exception as e:
                logger.error(f"解析CSV文件失败: {e}")
                raise HTTPException(status_code=500, detail=f"CSV文件解析失败: {str(e)}")
                
        elif file_name.endswith(('.xlsx', '.xls', '.et')):
            # Excel文件预览 - 使用超级稳定的多层降级策略
            try:
                df = None
                errors = []
                success_method = None

                # 策略1: 标准pandas读取（最快最简单）
                try:
                    if file_name.endswith('.xlsx'):
                        df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl', nrows=lines)
                    elif file_name.endswith('.xls'):
                        df = pd.read_excel(io.BytesIO(file_data), engine='xlrd', nrows=lines)
                    else:  # .et文件
                        df = pd.read_excel(io.BytesIO(file_data), engine='openpyxl', nrows=lines)
                    
                    success_method = "标准pandas读取"
                    logger.info(f"策略1成功: {success_method}")
                    
                except Exception as e1:
                    errors.append(f"标准读取: {str(e1)}")
                    logger.warning(f"策略1失败: {e1}")

                # 策略2: openpyxl直接读取（跳过pandas的一些检查）
                if df is None and file_name.endswith(('.xlsx', '.et')):
                    try:
                        import openpyxl
                        
                        # 使用read_only模式，更稳定
                        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
                        ws = wb.active
                        
                        # 读取数据
                        data = []
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i >= lines + 1:  # +1 因为包含标题行
                                break
                            data.append(row)
                        
                        wb.close()
                        
                        if data and len(data) > 0:
                            # 处理列名
                            cols = data[0] if data[0] else []
                            cols = [str(col) if col is not None else f'Column_{i}' for i, col in enumerate(cols)]
                            
                            # 处理数据行
                            rows = data[1:] if len(data) > 1 else []
                            clean_rows = []
                            
                            for row in rows:
                                row_list = list(row) if row else []
                                # 确保行长度与列数一致
                                if len(row_list) < len(cols):
                                    row_list.extend([None] * (len(cols) - len(row_list)))
                                elif len(row_list) > len(cols):
                                    row_list = row_list[:len(cols)]
                                clean_rows.append(row_list)
                            
                            df = pd.DataFrame(clean_rows, columns=cols)
                            success_method = "openpyxl只读模式"
                            logger.info(f"策略2成功: {success_method}")
                        
                    except Exception as e2:
                        errors.append(f"openpyxl只读: {str(e2)}")
                        logger.warning(f"策略2失败: {e2}")

                # 策略3: xlrd读取（适用于老版本Excel文件）
                if df is None and file_name.endswith(('.xls', '.xlsx', '.et')):
                    try:
                        import xlrd
                        
                        # 使用xlrd直接读取
                        workbook = xlrd.open_workbook(file_contents=file_data)
                        sheet = workbook.sheet_by_index(0)
                        
                        # 读取数据
                        data = []
                        max_rows = min(sheet.nrows, lines + 1)  # +1 包含标题行
                        
                        for row_idx in range(max_rows):
                            row_data = []
                            for col_idx in range(sheet.ncols):
                                cell = sheet.cell(row_idx, col_idx)
                                row_data.append(cell.value)
                            data.append(row_data)
                        
                        if data and len(data) > 0:
                            # 处理列名
                            cols = data[0] if data[0] else []
                            cols = [str(col) if col is not None else f'Column_{i}' for i, col in enumerate(cols)]
                            
                            # 处理数据行
                            rows = data[1:] if len(data) > 1 else []
                            clean_rows = []
                            
                            for row in rows:
                                row_list = list(row) if row else []
                                # 确保行长度与列数一致
                                if len(row_list) < len(cols):
                                    row_list.extend([None] * (len(cols) - len(row_list)))
                                elif len(row_list) > len(cols):
                                    row_list = row_list[:len(cols)]
                                clean_rows.append(row_list)
                            
                            df = pd.DataFrame(clean_rows, columns=cols)
                            success_method = "xlrd直接读取"
                            logger.info(f"策略3成功: {success_method}")
                        
                    except Exception as e3:
                        errors.append(f"xlrd直接读取: {str(e3)}")
                        logger.warning(f"策略3失败: {e3}")

                # 策略4: 尝试CSV格式读取（某些Excel文件可能是CSV格式）
                if df is None:
                    try:
                        # 尝试不同编码
                        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig', 'latin1']
                        for encoding in encodings:
                            try:
                                df = pd.read_csv(
                                    io.BytesIO(file_data),
                                    encoding=encoding,
                                    nrows=lines
                                )
                                success_method = f"CSV格式读取({encoding})"
                                logger.info(f"策略4成功: {success_method}")
                                break
                            except UnicodeDecodeError:
                                continue
                        
                    except Exception as e4:
                        errors.append(f"CSV格式读取: {str(e4)}")
                        logger.warning(f"策略4失败: {e4}")

                # 如果所有策略都失败
                if df is None:
                    error_detail = f"所有解析策略均失败: {'; '.join(errors)}"
                    if file_name.endswith('.et'):
                        error_detail += "\n\n建议: WPS .et格式兼容性有限，请将文件另存为 .xlsx 或 .csv 格式后重新上传。"
                    
                    logger.error(error_detail)
                    raise HTTPException(status_code=500, detail=f"文件解析失败。{error_detail}")
                
                # 数据清理和验证
                if df is not None:
                    # 清理空行和空列
                    df = df.dropna(how='all')  # 删除全空行
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]  # 删除未命名列
                    
                    # 限制数据量
                    if len(df) > lines:
                        df = df.head(lines)
                
                # 转换为JSON格式返回
                preview_data = df.to_dict('records')
                
                return {
                    "success": True,
                    "file_type": "excel",
                    "total_rows": len(preview_data),
                    "columns": list(df.columns),
                    "data": preview_data,
                    "message": f"显示前 {len(preview_data)} 行数据 (使用{success_method})",
                    "parse_method": success_method
                }
                
            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Excel文件预览发生未预期错误: {e}")
                raise HTTPException(status_code=500, detail=f"文件预览失败: {str(e)}")
        
        else:
            # 其他文件类型，返回原始文本内容
            try:
                # 尝试以文本形式读取
                encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
                content = None
                used_encoding = None
                
                for encoding in encodings:
                    try:
                        content = file_data.decode(encoding)
                        used_encoding = encoding
                        break
                    except UnicodeDecodeError:
                        continue
                
                if content is None:
                    # 如果无法解码为文本，返回二进制信息
                    return {
                        "success": True,
                        "file_type": "binary",
                        "file_size": len(file_data),
                        "message": "该文件为二进制文件，无法预览文本内容"
                    }
                
                # 限制预览行数
                lines_list = content.split('\n')[:lines]
                preview_content = '\n'.join(lines_list)
                
                return {
                    "success": True,
                    "file_type": "text",
                    "encoding": used_encoding,
                    "total_lines": len(lines_list),
                    "content": preview_content,
                    "message": f"显示前 {len(lines_list)} 行内容"
                }
                
            except Exception as e:
                logger.error(f"读取文件内容失败: {e}")
                raise HTTPException(status_code=500, detail=f"文件读取失败: {str(e)}")
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取文件预览失败: {e}")
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


@router.get("/datasets/{dataset_id}/download-original")
async def download_original_file(
    dataset_id: str,
    session: AsyncSession = Depends(get_async_session)
):
    """
    下载数据集的原文件
    
    Args:
        dataset_id: 数据集ID
        session: 数据库会话
        
    Returns:
        文件下载响应
    """
    try:
        # 查询数据集信息
        result = await session.execute(
            select(SysDataset).where(SysDataset.id == dataset_id)
        )
        dataset = result.scalar_one_or_none()
        
        if not dataset:
            raise HTTPException(status_code=404, detail="数据集不存在")
            
        if not dataset.original_file_path:
            raise HTTPException(status_code=404, detail="原文件路径不存在")
            
        # 从MinIO路径中提取对象名称
        object_name = dataset.original_file_path
        if '/' in object_name and object_name.count('/') > 1:
            parts = object_name.split('/', 1)
            if len(parts) > 1:
                object_name = parts[1]
        
        # 检查文件是否存在
        if not minio_client.file_exists(object_name):
            raise HTTPException(status_code=404, detail="原文件不存在")
            
        # 从MinIO下载文件
        file_data = minio_client.download_file(object_name)
        
        # 确定MIME类型
        content_type = "application/octet-stream"
        if dataset.name.lower().endswith('.csv'):
            content_type = "text/csv"
        elif dataset.name.lower().endswith(('.xlsx', '.xls')):
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif dataset.name.lower().endswith('.et'):
            content_type = "application/vnd.ms-excel"
        elif dataset.name.lower().endswith('.pdf'):
            content_type = "application/pdf"
        elif dataset.name.lower().endswith(('.docx', '.doc')):
            content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        elif dataset.name.lower().endswith(('.pptx', '.ppt')):
            content_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            
        # 处理文件名编码问题 - 使用URL编码
        import urllib.parse
        encoded_filename = urllib.parse.quote(dataset.name.encode('utf-8'))
            
        return Response(
            content=file_data,
            media_type=content_type,
            headers={
                "Content-Disposition": f'attachment; filename*=UTF-8\'\'{encoded_filename}',
                "Content-Length": str(len(file_data))
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载原文件失败: {e}")
        raise HTTPException(status_code=500, detail=f"下载失败: {str(e)}")